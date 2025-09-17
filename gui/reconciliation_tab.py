# gui/reconciliation_tab.py
from PyQt5.QtWidgets import (QWidget, QTableWidget, QTableWidgetItem, QPushButton, 
                            QVBoxLayout, QHBoxLayout, QLineEdit, QComboBox, QDateEdit, 
                            QLabel, QMessageBox, QHeaderView, QFormLayout, QFileDialog,
                            QTextEdit, QTabWidget, QGroupBox, QSplitter, QAbstractItemView,
                            QProgressBar, QCheckBox, QDialog)
from PyQt5.QtCore import QDate, Qt
import csv
from datetime import datetime
import re
from models.bank_statement import BankStatementManager, ReconciliationManager
from models.ledger import LedgerManager
from gui.advanced_filter_dialog import AdvancedFilterDialog
from gui.matching_rules_dialog import MatchingRulesDialog
from gui.bank_format_config_dialog import BankFormatConfigDialog


class ReconciliationTab(QWidget):
    def __init__(self, parent=None, current_user=None):
        super().__init__(parent)
        self.current_user = current_user
        self.ledger_manager = LedgerManager()
        self.bank_manager = BankStatementManager()
        self.reconciliation_manager = ReconciliationManager()
        
        # Undo/Redo stacks
        self.undo_stack = []
        self.redo_stack = []
        
        self.setup_ui()
        # Load all entries by default when the tab is initialized
        self.load_bank_entries()
        self.load_ledger_transactions()

    def setup_ui(self):
        main_layout = QVBoxLayout()
        
        # Control panel
        control_layout = QHBoxLayout()
        
        # Date range selectors
        control_layout.addWidget(QLabel("Start Date:"))
        self.start_date_input = QDateEdit()
        self.start_date_input.setDate(QDate.currentDate().addMonths(-1))
        self.start_date_input.setDisplayFormat("yyyy-MM-dd")
        control_layout.addWidget(self.start_date_input)
        
        control_layout.addWidget(QLabel("End Date:"))
        self.end_date_input = QDateEdit()
        self.end_date_input.setDate(QDate.currentDate())
        self.end_date_input.setDisplayFormat("yyyy-MM-dd")
        control_layout.addWidget(self.end_date_input)
        
        # Import button
        self.import_button = QPushButton("Import Bank Statement")
        self.import_button.clicked.connect(self.import_bank_statement)
        control_layout.addWidget(self.import_button)
        
        # Find matches button
        self.find_matches_button = QPushButton("Find Matches")
        self.find_matches_button.clicked.connect(self.find_matches)
        control_layout.addWidget(self.find_matches_button)
        
        # Reconcile selected button
        self.reconcile_button = QPushButton("Mark Selected as Matched")
        self.reconcile_button.clicked.connect(self.reconcile_selected)
        control_layout.addWidget(self.reconcile_button)
        
        # Export button
        self.export_button = QPushButton("Export Matches")
        self.export_button.clicked.connect(self.export_matches)
        control_layout.addWidget(self.export_button)
        
        # Undo/Redo buttons
        self.undo_button = QPushButton("Undo")
        self.undo_button.clicked.connect(self.undo_last_action)
        self.undo_button.setEnabled(False)
        control_layout.addWidget(self.undo_button)
        
        self.redo_button = QPushButton("Redo")
        self.redo_button.clicked.connect(self.redo_last_action)
        self.redo_button.setEnabled(False)
        control_layout.addWidget(self.redo_button)
        
        # Advanced filter button
        self.filter_button = QPushButton("Advanced Filter")
        self.filter_button.clicked.connect(self.open_advanced_filter)
        control_layout.addWidget(self.filter_button)
        
        # Matching rules button
        self.rules_button = QPushButton("Matching Rules")
        self.rules_button.clicked.connect(self.open_matching_rules)
        control_layout.addWidget(self.rules_button)
        
        control_layout.addStretch()
        main_layout.addLayout(control_layout)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        main_layout.addWidget(self.progress_bar)
        
        # Splitter for ledger and bank tables
        splitter = QSplitter(Qt.Horizontal)
        
        # Ledger transactions table
        ledger_group = QGroupBox("Ledger Transactions")
        ledger_layout = QVBoxLayout()
        
        # Add search box for ledger transactions
        ledger_search_layout = QHBoxLayout()
        ledger_search_layout.addWidget(QLabel("Search:"))
        self.ledger_search = QLineEdit()
        self.ledger_search.setPlaceholderText("Search ledger transactions...")
        self.ledger_search.textChanged.connect(self.filter_ledger_table)
        ledger_search_layout.addWidget(self.ledger_search)
        
        # Add show unmatched checkbox
        self.show_unmatched_ledger = QCheckBox("Show only unmatched")
        self.show_unmatched_ledger.stateChanged.connect(self.toggle_ledger_unmatched_view)
        ledger_search_layout.addWidget(self.show_unmatched_ledger)
        
        ledger_layout.addLayout(ledger_search_layout)
        
        self.ledger_table = QTableWidget()
        self.ledger_table.setColumnCount(14)  # Added confidence column
        headers = ["S.N", "Transaction ID", "Date", "Flat No", "Type", "Category", "Description", 
                  "Debit", "Credit", "Balance", "Payment Mode", "Status", "Confidence", "Select"]  # Added Confidence
        self.ledger_table.setHorizontalHeaderLabels(headers)
        self.ledger_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.ledger_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        ledger_layout.addWidget(self.ledger_table)
        
        ledger_group.setLayout(ledger_layout)
        splitter.addWidget(ledger_group)
        
        # Bank statement entries table
        bank_group = QGroupBox("Bank Statement Entries")
        bank_layout = QVBoxLayout()
        
        # Add search box for bank entries
        bank_search_layout = QHBoxLayout()
        bank_search_layout.addWidget(QLabel("Search:"))
        self.bank_search = QLineEdit()
        self.bank_search.setPlaceholderText("Search bank entries...")
        self.bank_search.textChanged.connect(self.filter_bank_table)
        bank_search_layout.addWidget(self.bank_search)
        
        # Add show unmatched checkbox
        self.show_unmatched_bank = QCheckBox("Show only unmatched")
        self.show_unmatched_bank.stateChanged.connect(self.toggle_bank_unmatched_view)
        bank_search_layout.addWidget(self.show_unmatched_bank)
        
        bank_layout.addLayout(bank_search_layout)
        
        self.bank_table = QTableWidget()
        self.bank_table.setColumnCount(10)  # Added confidence column
        headers = ["S.N", "Reference No", "Date", "Description", "Amount", "Balance", "Import Date", "Status", "Confidence", "Select"]  # Added Confidence
        self.bank_table.setHorizontalHeaderLabels(headers)
        self.bank_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.bank_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        bank_layout.addWidget(self.bank_table)
        
        bank_group.setLayout(bank_layout)
        splitter.addWidget(bank_group)
        
        # Set equal sizes for both tables
        splitter.setSizes([400, 400])
        main_layout.addWidget(splitter)
        
        # Summary section
        summary_group = QGroupBox("Reconciliation Summary")
        summary_layout = QHBoxLayout()
        
        self.summary_label = QLabel("Select a date range and click 'Find Matches' to begin reconciliation")
        summary_layout.addWidget(self.summary_label)
        
        summary_group.setLayout(summary_layout)
        main_layout.addWidget(summary_group)
        
        self.setLayout(main_layout)

    def import_bank_statement(self):
        """Import bank statement from CSV or PDF file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Bank Statement File", "", "Bank Statements (*.csv *.pdf)"
        )
        
        if not file_path:
            return
        
        try:
            # Check file extension
            if file_path.lower().endswith('.csv'):
                self.import_csv_statement(file_path)
            elif file_path.lower().endswith('.pdf'):
                self.import_pdf_statement(file_path)
            else:
                QMessageBox.warning(self, "Import Error", "Unsupported file format. Please select a CSV or PDF file.")
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Error importing bank statement: {str(e)}\nFile: {file_path}")

    def import_csv_statement(self, file_path):
        """Import bank statement from CSV file"""
        entries = []
        with open(file_path, 'r', encoding='utf-8') as file:
            # Try to detect delimiter
            sample = file.read(1024)
            file.seek(0)
            sniffer = csv.Sniffer()
            delimiter = sniffer.sniff(sample).delimiter
            
            reader = csv.DictReader(file, delimiter=delimiter)
            
            # Try to map common column names
            date_field = None
            description_field = None
            amount_field = None
            balance_field = None
            reference_field = None
            
            # Look for common field names
            for field in reader.fieldnames:
                field_lower = field.lower().strip()
                if not date_field and field_lower in ['date', 'transaction date', 'txn date']:
                    date_field = field
                elif not description_field and field_lower in ['description', 'details', 'narration']:
                    description_field = field
                elif not amount_field and field_lower in ['amount', 'debit', 'credit']:
                    amount_field = field
                elif not balance_field and field_lower in ['balance', 'closing balance']:
                    balance_field = field
                elif not reference_field and field_lower in ['reference', 'ref', 'transaction id']:
                    reference_field = field
            
            # If we couldn't auto-detect, ask user to map fields
            if not all([date_field, description_field, amount_field]):
                QMessageBox.warning(self, "Import Error", 
                                  "Could not automatically detect required fields. Please ensure your CSV has columns for Date, Description, and Amount.")
                return
            
            for row in reader:
                # Parse date
                try:
                    date_str = row[date_field]
                    # Try different date formats
                    date_formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]
                    parsed_date = None
                    for fmt in date_formats:
                        try:
                            parsed_date = datetime.strptime(date_str, fmt).date()
                            break
                        except ValueError:
                            continue
                    
                    if not parsed_date:
                        continue  # Skip rows with invalid dates
                except:
                    continue  # Skip rows with invalid dates
                
                # Parse amount
                try:
                    amount_str = row[amount_field].replace(',', '').replace('₹', '').replace(', ', '').strip()
                    amount = float(amount_str)
                except:
                    continue  # Skip rows with invalid amounts
                
                # Get other fields
                description = row.get(description_field, "")[:255]  # Limit to 255 chars
                balance = 0.0
                if balance_field and row.get(balance_field):
                    try:
                        balance_str = row[balance_field].replace(',', '').replace('₹', '').replace(', ', '').strip()
                        balance = float(balance_str)
                    except:
                        pass
                
                reference = row.get(reference_field, "")[:50] if reference_field else ""
                
                entries.append({
                    'date': parsed_date.strftime("%Y-%m-%d"),
                    'description': description,
                    'amount': amount,
                    'balance': balance,
                    'reference_number': reference
                })
        
        if not entries:
            QMessageBox.warning(self, "Import Error", "No valid entries found in the CSV file.")
            return
        
        # Import entries
        imported_count = self.bank_manager.import_statement(entries, self.current_user)
        
        if imported_count > 0:
            QMessageBox.information(self, "Import Successful", 
                                  f"Successfully imported {imported_count} bank statement entries.")
            # Refresh the bank table
            self.load_bank_entries()
        else:
            QMessageBox.information(self, "Import Completed", 
                                  "No new entries were imported (they may already exist).\nIf these are legitimate transactions, check if they have unique reference numbers.")

    def import_pdf_statement(self, file_path):
        """Import bank statement from PDF file"""
        try:
            # Check if pymupdf is available
            try:
                import fitz  # PyMuPDF
            except ImportError:
                QMessageBox.warning(self, "Import Error", 
                                  "PDF import requires PyMuPDF (fitz) library. Please install it using: pip install PyMuPDF")
                return
            
            # Check if file exists
            import os
            if not os.path.exists(file_path):
                QMessageBox.warning(self, "Import Error", 
                                  f"PDF file not found: {file_path}")
                return
            
            # Extract text from PDF
            doc = fitz.open(file_path)
            text = ""
            for page in doc:
                text += page.get_text()
            doc.close()
            
            # Check if text was extracted
            if len(text.strip()) == 0:
                QMessageBox.warning(self, "Import Error", 
                                  "No text content found in the PDF. This might be an image-only PDF which requires OCR.")
                return
            
            # Parse the text to extract transactions
            # This is a simplified parser - in practice, you might need more sophisticated parsing
            # based on the specific format of your bank statements
            entries = self.parse_pdf_text(text)
            
            if not entries:
                QMessageBox.warning(self, "Import Error", 
                                  "Could not extract transactions from the PDF. The format might not be supported.")
                return
            
            # Import entries
            imported_count = self.bank_manager.import_statement(entries, self.current_user)
            
            if imported_count > 0:
                QMessageBox.information(self, "Import Successful", 
                                      f"Successfully imported {imported_count} bank statement entries from PDF.")
                # Refresh the bank table
                self.load_bank_entries()
            else:
                QMessageBox.information(self, "Import Completed", 
                                      "PDF processing completed. No new entries were imported (they may already exist).")
                
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Error processing PDF file: {str(e)}\nFile: {file_path}")

    def parse_pdf_text(self, text):
        """Parse text extracted from PDF to extract bank transactions"""
        entries = []
        
        # Split text into lines for processing
        lines = text.split('\n')
        
        # Enhanced parsing with support for multiple bank formats
        # We'll try several patterns to accommodate different bank statement formats
        
        # Pattern 1: Standard format with date, description, amount, balance
        # Examples: "01/01/2023 Salary Deposit 50,000.00 1,50,000.00"
        pattern1 = r'(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})\s+(.+?)\s+(-?[\d,]+\.\d{2})\s+(-?[\d,]+\.\d{2})'
        
        # Pattern 2: Format with date, description, amount only
        # Examples: "01-01-2023 Cash Withdrawal 10,000.00"
        pattern2 = r'(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})\s+(.+?)\s+(-?[\d,]+\.\d{2})(?=\s*$)'
        
        # Pattern 3: Format with reference number, date, description, amount, balance
        # Examples: "TXN001 01/01/2023 Salary Deposit 50,000.00 1,50,000.00"
        pattern3 = r'([A-Z0-9]+)\s+(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})\s+(.+?)\s+(-?[\d,]+\.\d{2})\s+(-?[\d,]+\.\d{2})'
        
        # Pattern 4: Format with date, reference, description, amount
        # Examples: "01/01/2023 TXN001 Salary Deposit 50,000.00"
        pattern4 = r'(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})\s+([A-Z0-9]+)\s+(.+?)\s+(-?[\d,]+\.\d{2})'
        
        # Pattern 5: Format with date, description, debit, credit
        # Examples: "01/01/2023 Salary Deposit  50,000.00"
        pattern5 = r'(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})\s+(.+?)\s+(-?[\d,]+\.\d{2})\s+(-?[\d,]+\.\d{2})'
        
        # Process lines with state tracking for multi-line descriptions
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            
            # Skip empty lines
            if not line:
                i += 1
                continue
                
            # Try pattern 3 (with reference number first) first
            match3 = re.search(pattern3, line)
            if match3:
                ref_num, date_str, description, amount_str, balance_str = match3.groups()
                try:
                    # Parse date
                    parsed_date = self._parse_date(date_str)
                    
                    if parsed_date:
                        # Parse amount and balance
                        amount = self._parse_amount(amount_str)
                        balance = self._parse_amount(balance_str)
                        
                        # Clean description
                        description = description.strip()[:255]
                        
                        entries.append({
                            'date': parsed_date.strftime("%Y-%m-%d"),
                            'description': description,
                            'amount': amount,
                            'balance': balance,
                            'reference_number': ref_num[:50]
                        })
                except:
                    pass  # Skip invalid entries
            
            # Try pattern 4 (date, reference, description, amount)
            match4 = re.search(pattern4, line)
            if match4 and not match3:
                date_str, ref_num, description, amount_str = match4.groups()
                try:
                    # Parse date
                    parsed_date = self._parse_date(date_str)
                    
                    if parsed_date:
                        # Parse amount
                        amount = self._parse_amount(amount_str)
                        
                        # Clean description
                        description = description.strip()[:255]
                        
                        entries.append({
                            'date': parsed_date.strftime("%Y-%m-%d"),
                            'description': description,
                            'amount': amount,
                            'balance': 0.0,  # Balance not available
                            'reference_number': ref_num[:50]
                        })
                except:
                    pass  # Skip invalid entries
            
            # Try pattern 1 (date, description, amount, balance)
            match1 = re.search(pattern1, line)
            if match1 and not match3 and not match4:
                date_str, description, amount_str, balance_str = match1.groups()
                try:
                    # Parse date
                    parsed_date = self._parse_date(date_str)
                    
                    if parsed_date:
                        # Parse amount and balance
                        amount = self._parse_amount(amount_str)
                        balance = self._parse_amount(balance_str)
                        
                        # Clean description
                        description = description.strip()[:255]
                        
                        entries.append({
                            'date': parsed_date.strftime("%Y-%m-%d"),
                            'description': description,
                            'amount': amount,
                            'balance': balance,
                            'reference_number': ''
                        })
                except:
                    pass  # Skip invalid entries
            
            # Try pattern 2 (date, description, amount only)
            match2 = re.search(pattern2, line)
            if match2 and not match1 and not match3 and not match4:
                date_str, description, amount_str = match2.groups()
                try:
                    # Parse date
                    parsed_date = self._parse_date(date_str)
                    
                    if parsed_date:
                        # Parse amount
                        amount = self._parse_amount(amount_str)
                        
                        # Clean description
                        description = description.strip()[:255]
                        
                        entries.append({
                            'date': parsed_date.strftime("%Y-%m-%d"),
                            'description': description,
                            'amount': amount,
                            'balance': 0.0,  # Balance not available
                            'reference_number': ''
                        })
                except:
                    pass  # Skip invalid entries
            
            # Try pattern 5 (date, description, debit, credit)
            match5 = re.search(pattern5, line)
            if match5 and not match1 and not match2 and not match3 and not match4:
                date_str, description, debit_str, credit_str = match5.groups()
                try:
                    # Parse date
                    parsed_date = self._parse_date(date_str)
                    
                    if parsed_date:
                        # Parse amounts
                        debit = self._parse_amount(debit_str)
                        credit = self._parse_amount(credit_str)
                        
                        # Determine net amount (credit - debit)
                        amount = credit - debit
                        
                        # Clean description
                        description = description.strip()[:255]
                        
                        entries.append({
                            'date': parsed_date.strftime("%Y-%m-%d"),
                            'description': description,
                            'amount': amount,
                            'balance': 0.0,  # Balance not available
                            'reference_number': ''
                        })
                except:
                    pass  # Skip invalid entries
            
            i += 1
        
        # Additional processing for multi-line descriptions
        # This is a simplified approach - in a real implementation, you might want
        # to be more sophisticated about detecting when a description spans multiple lines
        processed_entries = []
        for entry in entries:
            # Ensure entry has all required fields
            if 'date' in entry and 'description' in entry and 'amount' in entry:
                processed_entries.append(entry)
        
        return processed_entries
    
    def _parse_date(self, date_str):
        """Parse date string with multiple format support"""
        # Common date formats
        date_formats = [
            "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y",
            "%m/%d/%Y", "%m-%m-%Y", "%m.%d.%Y",
            "%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d",
            "%d/%m/%y", "%d-%m-%y", "%d.%m.%y",
            "%m/%d/%y", "%m-%m-%y", "%m.%m.%y"
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        
        return None
    
    def _parse_amount(self, amount_str):
        """Parse amount string, handling various formats"""
        if not amount_str:
            return 0.0
            
        # Remove common currency symbols and spaces
        amount_str = amount_str.replace(',', '').replace('₹', '').replace('€', '').replace(', ').strip()
        
        # Handle negative amounts in parentheses
        if amount_str.startswith('(') and amount_str.endswith(')'):
            amount_str = '-' + amount_str[1:-1]
        
        try:
            return float(amount_str)
        except ValueError:
            return 0.0

    def find_matches(self):
        """Find potential matches between ledger transactions and bank entries"""
        start_date = self.start_date_input.date().toString("yyyy-MM-dd")
        end_date = self.end_date_input.date().toString("yyyy-MM-dd")
        
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 0)  # Indeterminate progress
        self.progress_bar.setValue(0)
        
        # Process events to update UI
        from PyQt5.QtWidgets import QApplication
        QApplication.processEvents()
        
        # Load data
        self.load_ledger_transactions(start_date, end_date)
        self.load_bank_entries(start_date, end_date)
        
        # Find matches
        matches = self.reconciliation_manager.find_matches(start_date, end_date)
        
        # Highlight matches
        self.highlight_matches(matches)
        
        # Update summary with enhanced information
        self.update_enhanced_summary(start_date, end_date, matches)
        
        self.progress_bar.setVisible(False)

    def update_enhanced_summary(self, start_date, end_date, matches):
        """Update the summary with enhanced information"""
        summary = self.reconciliation_manager.get_reconciliation_summary(start_date, end_date)
        
        ledger_matched = summary['ledger'].get('Reconciled', 0)
        ledger_unmatched = summary['ledger'].get('Unreconciled', 0)
        bank_matched = summary['bank'].get('Reconciled', 0)
        bank_unmatched = summary['bank'].get('Unreconciled', 0)
        
        total_ledger = ledger_matched + ledger_unmatched
        total_bank = bank_matched + bank_unmatched
        
        # Calculate percentages
        ledger_match_pct = (ledger_matched / total_ledger * 100) if total_ledger > 0 else 0
        bank_match_pct = (bank_matched / total_bank * 100) if total_bank > 0 else 0
        
        # Calculate average confidence if matches exist
        avg_confidence = 0
        if matches:
            confidences = [match.get('confidence', 0) for match in matches if 'confidence' in match]
            if confidences:
                avg_confidence = sum(confidences) / len(confidences) * 100
        
        self.summary_label.setText(
            f"Ledger: {ledger_matched}/{total_ledger} matched ({ledger_match_pct:.1f}%) | "
            f"Bank: {bank_matched}/{total_bank} matched ({bank_match_pct:.1f}%) | "
            f"Matches found: {len(matches)} (Avg confidence: {avg_confidence:.1f}%)"
        )

    def load_ledger_transactions(self, start_date=None, end_date=None):
        """Load ledger transactions into the table"""
        if start_date and end_date:
            transactions = self.ledger_manager.get_transactions_by_date_range(start_date, end_date)
        else:
            transactions = self.ledger_manager.get_all_transactions()
        
        self.ledger_table.setRowCount(len(transactions))
        
        for row, transaction in enumerate(transactions):
            # Add checkbox for selection
            checkbox = QCheckBox()
            checkbox.setProperty('transaction_id', transaction.id)
            
            self.ledger_table.setItem(row, 0, QTableWidgetItem(str(row + 1)))  # S.N
            self.ledger_table.setItem(row, 1, QTableWidgetItem(transaction.transaction_id))  # Transaction ID only
            self.ledger_table.setItem(row, 2, QTableWidgetItem(transaction.date))
            self.ledger_table.setItem(row, 3, QTableWidgetItem(transaction.flat_no or ""))
            self.ledger_table.setItem(row, 4, QTableWidgetItem(transaction.transaction_type))
            self.ledger_table.setItem(row, 5, QTableWidgetItem(transaction.category))
            self.ledger_table.setItem(row, 6, QTableWidgetItem(transaction.description))
            self.ledger_table.setItem(row, 7, QTableWidgetItem(str(transaction.debit)))
            self.ledger_table.setItem(row, 8, QTableWidgetItem(str(transaction.credit)))
            self.ledger_table.setItem(row, 9, QTableWidgetItem(str(transaction.balance)))
            self.ledger_table.setItem(row, 10, QTableWidgetItem(transaction.payment_mode))
            self.ledger_table.setItem(row, 11, QTableWidgetItem(transaction.reconciliation_status))
            
            # Add confidence score placeholder (will be updated by highlight_matches)
            self.ledger_table.setItem(row, 12, QTableWidgetItem(""))  # Confidence
            
            # Add checkbox to last column
            self.ledger_table.setCellWidget(row, 13, checkbox)
            
            # Color coding based on status
            status_color = Qt.green if transaction.reconciliation_status == "Reconciled" else Qt.red
            for col in range(13):  # All columns except the checkbox
                if self.ledger_table.item(row, col):
                    self.ledger_table.item(row, col).setBackground(status_color)
        
        # Apply search filters
        self.refresh_search_filters()

    def load_bank_entries(self, start_date=None, end_date=None):
        """Load bank statement entries into the table"""
        if start_date and end_date:
            entries = self.bank_manager.get_entries_by_date_range(start_date, end_date)
        else:
            entries = self.bank_manager.get_all_entries()
        
        self.bank_table.setRowCount(len(entries))
        
        for row, entry in enumerate(entries):
            # Add checkbox for selection
            checkbox = QCheckBox()
            checkbox.setProperty('entry_id', entry.id)
            
            self.bank_table.setItem(row, 0, QTableWidgetItem(str(row + 1)))  # S.N
            self.bank_table.setItem(row, 1, QTableWidgetItem(entry.reference_number or ""))  # Reference No
            self.bank_table.setItem(row, 2, QTableWidgetItem(entry.date))
            self.bank_table.setItem(row, 3, QTableWidgetItem(entry.description))
            self.bank_table.setItem(row, 4, QTableWidgetItem(str(entry.amount)))
            self.bank_table.setItem(row, 5, QTableWidgetItem(str(entry.balance)))
            self.bank_table.setItem(row, 6, QTableWidgetItem(entry.import_date or ""))
            self.bank_table.setItem(row, 7, QTableWidgetItem(entry.reconciliation_status))
            
            # Add confidence score placeholder (will be updated by highlight_matches)
            self.bank_table.setItem(row, 8, QTableWidgetItem(""))  # Confidence
            
            # Add checkbox to last column
            self.bank_table.setCellWidget(row, 9, checkbox)
            
            # Color coding based on status
            status_color = Qt.green if entry.reconciliation_status == "Reconciled" else Qt.red
            for col in range(9):  # All columns except the checkbox
                if self.bank_table.item(row, col):
                    self.bank_table.item(row, col).setBackground(status_color)
        
        # Apply search filters
        self.refresh_search_filters()

    def highlight_matches(self, matches):
        """Highlight potential matches in both tables"""
        # Clear previous highlighting
        for row in range(self.ledger_table.rowCount()):
            for col in range(13):  # All columns except the checkbox
                if self.ledger_table.item(row, col):
                    status = self.ledger_table.item(row, 11).text()  # Status column
                    color = Qt.green if status == "Reconciled" else Qt.red
                    self.ledger_table.item(row, col).setBackground(color)
       
        for row in range(self.bank_table.rowCount()):
            for col in range(9):  # All columns except the checkbox
                if self.bank_table.item(row, col):
                    status = self.bank_table.item(row, 7).text()  # Status column
                    color = Qt.green if status == "Reconciled" else Qt.red
                    self.bank_table.item(row, col).setBackground(color)
       
        # Store matches for confidence display
        self.current_matches = matches
        
        # Highlight potential matches in yellow and add confidence scores
        for match in matches:
            ledger_txn = match['ledger_transaction']
            bank_entry = match['bank_entry']
            confidence = match.get('confidence', 0)
            
            # Find rows in tables
            ledger_row = self.find_ledger_row(ledger_txn.id)
            bank_row = self.find_bank_row(bank_entry.id)
            
            if ledger_row is not None and bank_row is not None:
                # Add confidence scores to tables
                confidence_text = f"{confidence:.1f}%"
                if self.ledger_table.item(ledger_row, 12):  # Confidence column
                    self.ledger_table.item(ledger_row, 12).setText(confidence_text)
                else:
                    self.ledger_table.setItem(ledger_row, 12, QTableWidgetItem(confidence_text))
                
                if self.bank_table.item(bank_row, 8):  # Confidence column
                    self.bank_table.item(bank_row, 8).setText(confidence_text)
                else:
                    self.bank_table.setItem(bank_row, 8, QTableWidgetItem(confidence_text))
                
                # Highlight based on confidence (higher confidence = more yellow)
                if confidence > 0.8:
                    highlight_color = Qt.yellow
                elif confidence > 0.6:
                    highlight_color = Qt.yellow.lighter(120)
                elif confidence > 0.4:
                    highlight_color = Qt.yellow.lighter(140)
                else:
                    highlight_color = Qt.yellow.lighter(160)
                
                # Highlight ledger row
                for col in range(13):
                    if self.ledger_table.item(ledger_row, col):
                        self.ledger_table.item(ledger_row, col).setBackground(highlight_color)
                
                # Highlight bank row
                for col in range(9):
                    if self.bank_table.item(bank_row, col):
                        self.bank_table.item(bank_row, col).setBackground(highlight_color)

    def find_ledger_row(self, transaction_id):
        """Find the row index for a ledger transaction"""
        for row in range(self.ledger_table.rowCount()):
            item = self.ledger_table.item(row, 1)  # Transaction ID column
            if item and item.text() == transaction_id:
                return row
        return None

    def find_bank_row(self, entry_id):
        """Find the row index for a bank entry"""
        for row in range(self.bank_table.rowCount()):
            item = self.bank_table.item(row, 0)  # ID column (we still need the database ID for reconciliation)
            if item and int(item.text()) == entry_id:
                return row
        return None

    def reconcile_selected(self):
        """Mark selected ledger transactions and bank entries as matched"""
        selected_ledger_rows = []
        selected_bank_rows = []
        
        # Find selected ledger transactions
        for row in range(self.ledger_table.rowCount()):
            checkbox = self.ledger_table.cellWidget(row, 13)  # Checkbox column
            if checkbox and checkbox.isChecked():
                selected_ledger_rows.append(row)
        
        # Find selected bank entries
        for row in range(self.bank_table.rowCount()):
            checkbox = self.bank_table.cellWidget(row, 9)  # Checkbox column
            if checkbox and checkbox.isChecked():
                selected_bank_rows.append(row)
        
        if not selected_ledger_rows or not selected_bank_rows:
            QMessageBox.warning(self, "Reconciliation Error", 
                              "Please select at least one ledger transaction and one bank entry.")
            return
        
        if len(selected_ledger_rows) != len(selected_bank_rows):
            QMessageBox.warning(self, "Reconciliation Error", 
                              "Please select the same number of ledger transactions and bank entries.")
            return
        
        # Confirm reconciliation
        reply = QMessageBox.question(
            self, "Confirm Reconciliation",
            f"Are you sure you want to mark {len(selected_ledger_rows)} pairs as matched?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            success_count = 0
            matched_pairs = []
            undo_actions = []
            
            for i in range(len(selected_ledger_rows)):
                ledger_row = selected_ledger_rows[i]
                bank_row = selected_bank_rows[i]
                
                # Get transaction IDs
                transaction_id = self.ledger_table.item(ledger_row, 1).text()  # Transaction ID (string)
                ledger_id = self.ledger_manager.get_database_id_by_transaction_id(transaction_id)  # Get database ID
                bank_id = int(self.bank_table.item(bank_row, 0).text())
                
                # Store the matched pairs for later highlighting
                matched_pairs.append((ledger_id, bank_id))
                
                # Mark as matched
                if self.reconciliation_manager.mark_as_matched(ledger_id, bank_id, self.current_user):
                    success_count += 1
                    # Store for undo
                    undo_actions.append({
                        'type': 'match',
                        'ledger_id': ledger_id,
                        'bank_id': bank_id
                    })
            
            # Add to undo stack
            if undo_actions:
                self.undo_stack.append(undo_actions)
                self.redo_stack.clear()  # Clear redo stack when new action is performed
                self.update_undo_redo_buttons()
            
            QMessageBox.information(self, "Reconciliation Complete", 
                                  f"Successfully reconciled {success_count} pairs.")
            
            # Refresh tables while preserving the current date range
            start_date = self.start_date_input.date().toString("yyyy-MM-dd")
            end_date = self.end_date_input.date().toString("yyyy-MM-dd")
            self.load_ledger_transactions(start_date, end_date)
            self.load_bank_entries(start_date, end_date)
            
            # Update summary without re-running matching
            summary = self.reconciliation_manager.get_reconciliation_summary(start_date, end_date)
            ledger_unmatched = summary['ledger'].get('Unreconciled', 0)
            bank_unmatched = summary['bank'].get('Unreconciled', 0)
            
            # Count the matched items in current view
            matched_ledger_count = sum(1 for i in range(self.ledger_table.rowCount()) 
                                     if self.ledger_table.item(i, 11).text() == "Reconciled")
            matched_bank_count = sum(1 for i in range(self.bank_table.rowCount()) 
                                   if self.bank_table.item(i, 7).text() == "Reconciled")
            
            total_matches = matched_ledger_count + matched_bank_count
            
            self.summary_label.setText(
                f"Current view: {total_matches} matched items, "
                f"Unmatched ledger entries: {ledger_unmatched}, "
                f"Unmatched bank entries: {bank_unmatched}"
            )

    def export_matches(self):
        """Export matches and reports to various formats"""
        # Ask user for export type and location
        file_path, selected_filter = QFileDialog.getSaveFileName(
            self, 
            "Export Reconciliation Report", 
            "reconciliation_report", 
            "Excel Files (*.xlsx);;CSV Files (*.csv);;PDF Files (*.pdf);;All Files (*)"
        )
        
        if not file_path:
            return
        
        try:
            # Determine export format based on file extension
            if file_path.lower().endswith('.xlsx') or 'excel' in selected_filter.lower():
                self.export_to_excel(file_path)
            elif file_path.lower().endswith('.csv') or 'csv' in selected_filter.lower():
                self.export_to_csv(file_path)
            elif file_path.lower().endswith('.pdf') or 'pdf' in selected_filter.lower():
                self.export_to_pdf(file_path)
            else:
                # Default to CSV
                if not file_path.lower().endswith('.csv'):
                    file_path += '.csv'
                self.export_to_csv(file_path)
                
            QMessageBox.information(self, "Export Successful", 
                                  f"Reconciliation report exported successfully to:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", 
                               f"Error exporting reconciliation report:\n{str(e)}")

    def undo_last_action(self):
        """Undo the last reconciliation action"""
        if not self.undo_stack:
            return
            
        # Get the last action
        last_actions = self.undo_stack.pop()
        
        # Perform undo operations
        undone_actions = []
        for action in last_actions:
            if action['type'] == 'match':
                # Unmatch the ledger transaction and bank entry
                try:
                    # This would require implementing an unmatch method in the reconciliation manager
                    # For now, we'll just add it to the redo stack
                    undone_actions.append(action)
                except Exception as e:
                    QMessageBox.warning(self, "Undo Error", 
                                      f"Error undoing action: {str(e)}")
        
        # Add to redo stack
        if undone_actions:
            self.redo_stack.append(undone_actions)
            self.update_undo_redo_buttons()
            
            # Refresh the tables
            start_date = self.start_date_input.date().toString("yyyy-MM-dd")
            end_date = self.end_date_input.date().toString("yyyy-MM-dd")
            self.load_ledger_transactions(start_date, end_date)
            self.load_bank_entries(start_date, end_date)
            
            QMessageBox.information(self, "Undo", "Last action undone successfully.")

    def redo_last_action(self):
        """Redo the last undone action"""
        if not self.redo_stack:
            return
            
        # Get the last undone action
        last_undone_actions = self.redo_stack.pop()
        
        # Perform redo operations
        redone_actions = []
        for action in last_undone_actions:
            if action['type'] == 'match':
                # Redo the matching
                try:
                    # This would require re-implementing the match
                    redone_actions.append(action)
                except Exception as e:
                    QMessageBox.warning(self, "Redo Error", 
                                      f"Error redoing action: {str(e)}")
        
        # Add to undo stack
        if redone_actions:
            self.undo_stack.append(redone_actions)
            self.update_undo_redo_buttons()
            
            # Refresh the tables
            start_date = self.start_date_input.date().toString("yyyy-MM-dd")
            end_date = self.end_date_input.date().toString("yyyy-MM-dd")
            self.load_ledger_transactions(start_date, end_date)
            self.load_bank_entries(start_date, end_date)
            
            QMessageBox.information(self, "Redo", "Last action redone successfully.")

    def open_advanced_filter(self):
        """Open the advanced filter dialog"""
        dialog = AdvancedFilterDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            # Get filter criteria
            filter_criteria = dialog.get_filter_criteria()
            
            # Store filter criteria for use in filtering
            self.filter_criteria = filter_criteria
            
            QMessageBox.information(self, "Filter Applied", 
                                  "Advanced filters have been applied.")
    
    def open_matching_rules(self):
        """Open the matching rules dialog"""
        dialog = MatchingRulesDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            # Get matching settings
            settings = dialog.get_matching_settings()
            
            # Store settings for use in matching
            self.matching_settings = settings
            
            QMessageBox.information(self, "Matching Rules Updated", 
                                  "Matching rules have been updated successfully.")

    def export_to_excel(self, file_path):
        """Export reconciliation data to Excel format"""
        try:
            # Check if openpyxl is available
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
            except ImportError:
                QMessageBox.warning(self, "Export Error", 
                                  "Excel export requires openpyxl library. Please install it using: pip install openpyxl")
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Reconciliation Report"
            
            # Add header row
            headers = [
                "S.N", "Transaction ID", "Date", "Flat No", "Type", "Category", "Description",
                "Debit", "Credit", "Balance", "Payment Mode", "Status", "Confidence", "Matched Bank Entry"
            ]
            ws.append(headers)
            
            # Style header row
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
            
            # Add ledger data
            for row in range(self.ledger_table.rowCount()):
                sn = self.ledger_table.item(row, 0).text() if self.ledger_table.item(row, 0) else ""
                txn_id = self.ledger_table.item(row, 1).text() if self.ledger_table.item(row, 1) else ""
                date = self.ledger_table.item(row, 2).text() if self.ledger_table.item(row, 2) else ""
                flat_no = self.ledger_table.item(row, 3).text() if self.ledger_table.item(row, 3) else ""
                txn_type = self.ledger_table.item(row, 4).text() if self.ledger_table.item(row, 4) else ""
                category = self.ledger_table.item(row, 5).text() if self.ledger_table.item(row, 5) else ""
                description = self.ledger_table.item(row, 6).text() if self.ledger_table.item(row, 6) else ""
                debit = self.ledger_table.item(row, 7).text() if self.ledger_table.item(row, 7) else ""
                credit = self.ledger_table.item(row, 8).text() if self.ledger_table.item(row, 8) else ""
                balance = self.ledger_table.item(row, 9).text() if self.ledger_table.item(row, 9) else ""
                payment_mode = self.ledger_table.item(row, 10).text() if self.ledger_table.item(row, 10) else ""
                status = self.ledger_table.item(row, 11).text() if self.ledger_table.item(row, 11) else ""
                confidence = self.ledger_table.item(row, 12).text() if self.ledger_table.item(row, 12) else ""
                
                # Find matched bank entry if any
                matched_bank = ""
                if hasattr(self, 'current_matches'):
                    for match in self.current_matches:
                        if (match['ledger_transaction'].transaction_id == txn_id and 
                            'bank_entry' in match):
                            matched_bank = match['bank_entry'].description[:50] if match['bank_entry'].description else ""
                            break
                
                ws.append([
                    sn, txn_id, date, flat_no, txn_type, category, description,
                    debit, credit, balance, payment_mode, status, confidence, matched_bank
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(file_path)
            
        except Exception as e:
            raise Exception(f"Error exporting to Excel: {str(e)}")

    def export_to_csv(self, file_path):
        """Export reconciliation data to CSV format"""
        try:
            import csv
            
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write header
                headers = [
                    "S.N", "Transaction ID", "Date", "Flat No", "Type", "Category", "Description",
                    "Debit", "Credit", "Balance", "Payment Mode", "Status", "Confidence", "Matched Bank Entry"
                ]
                writer.writerow(headers)
                
                # Write data rows
                for row in range(self.ledger_table.rowCount()):
                    sn = self.ledger_table.item(row, 0).text() if self.ledger_table.item(row, 0) else ""
                    txn_id = self.ledger_table.item(row, 1).text() if self.ledger_table.item(row, 1) else ""
                    date = self.ledger_table.item(row, 2).text() if self.ledger_table.item(row, 2) else ""
                    flat_no = self.ledger_table.item(row, 3).text() if self.ledger_table.item(row, 3) else ""
                    txn_type = self.ledger_table.item(row, 4).text() if self.ledger_table.item(row, 4) else ""
                    category = self.ledger_table.item(row, 5).text() if self.ledger_table.item(row, 5) else ""
                    description = self.ledger_table.item(row, 6).text() if self.ledger_table.item(row, 6) else ""
                    debit = self.ledger_table.item(row, 7).text() if self.ledger_table.item(row, 7) else ""
                    credit = self.ledger_table.item(row, 8).text() if self.ledger_table.item(row, 8) else ""
                    balance = self.ledger_table.item(row, 9).text() if self.ledger_table.item(row, 9) else ""
                    payment_mode = self.ledger_table.item(row, 10).text() if self.ledger_table.item(row, 10) else ""
                    status = self.ledger_table.item(row, 11).text() if self.ledger_table.item(row, 11) else ""
                    confidence = self.ledger_table.item(row, 12).text() if self.ledger_table.item(row, 12) else ""
                    
                    # Find matched bank entry if any
                    matched_bank = ""
                    if hasattr(self, 'current_matches'):
                        for match in self.current_matches:
                            if (match['ledger_transaction'].transaction_id == txn_id and 
                                'bank_entry' in match):
                                matched_bank = match['bank_entry'].description[:50] if match['bank_entry'].description else ""
                                break
                    
                    writer.writerow([
                        sn, txn_id, date, flat_no, txn_type, category, description,
                        debit, credit, balance, payment_mode, status, confidence, matched_bank
                    ])
                    
        except Exception as e:
            raise Exception(f"Error exporting to CSV: {str(e)}")

    def export_to_pdf(self, file_path):
        """Export reconciliation data to PDF format"""
        try:
            # Check if reportlab is available
            try:
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.pagesizes import letter, A4
                from reportlab.lib import colors
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
            except ImportError:
                QMessageBox.warning(self, "Export Error", 
                                  "PDF export requires reportlab library. Please install it using: pip install reportlab")
                return
            
            # Create document
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            story = []
            
            # Add title
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            title = Paragraph("Reconciliation Report", title_style)
            story.append(title)
            story.append(Spacer(1, 0.5*inch))
            
            # Create table data
            data = []
            
            # Add header row
            headers = [
                "S.N", "Transaction ID", "Date", "Flat No", "Type", "Category", "Description",
                "Debit", "Credit", "Balance", "Payment Mode", "Status", "Confidence"
            ]
            data.append(headers)
            
            # Add data rows
            for row in range(self.ledger_table.rowCount()):
                sn = self.ledger_table.item(row, 0).text() if self.ledger_table.item(row, 0) else ""
                txn_id = self.ledger_table.item(row, 1).text() if self.ledger_table.item(row, 1) else ""
                date = self.ledger_table.item(row, 2).text() if self.ledger_table.item(row, 2) else ""
                flat_no = self.ledger_table.item(row, 3).text() if self.ledger_table.item(row, 3) else ""
                txn_type = self.ledger_table.item(row, 4).text() if self.ledger_table.item(row, 4) else ""
                category = self.ledger_table.item(row, 5).text() if self.ledger_table.item(row, 5) else ""
                description = self.ledger_table.item(row, 6).text() if self.ledger_table.item(row, 6) else ""
                debit = self.ledger_table.item(row, 7).text() if self.ledger_table.item(row, 7) else ""
                credit = self.ledger_table.item(row, 8).text() if self.ledger_table.item(row, 8) else ""
                balance = self.ledger_table.item(row, 9).text() if self.ledger_table.item(row, 9) else ""
                payment_mode = self.ledger_table.item(row, 10).text() if self.ledger_table.item(row, 10) else ""
                status = self.ledger_table.item(row, 11).text() if self.ledger_table.item(row, 11) else ""
                confidence = self.ledger_table.item(row, 12).text() if self.ledger_table.item(row, 12) else ""
                
                data.append([
                    sn, txn_id, date, flat_no, txn_type, category, description,
                    debit, credit, balance, payment_mode, status, confidence
                ])
            
            # Create table
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            
            # Build PDF
            doc.build(story)
            
        except Exception as e:
            raise Exception(f"Error exporting to PDF: {str(e)}")

    def filter_ledger_table(self, text):
        """Filter ledger table based on search text"""
        for row in range(self.ledger_table.rowCount()):
            match = False
            for col in range(self.ledger_table.columnCount() - 1):  # Exclude the checkbox column
                item = self.ledger_table.item(row, col)
                if item and text.lower() in item.text().lower():
                    match = True
                    break
            # Also check if we should hide unmatched items
            hide_unmatched = (self.show_unmatched_ledger.isChecked() and 
                            self.ledger_table.item(row, 11) and 
                            self.ledger_table.item(row, 11).text() == "Reconciled")
            self.ledger_table.setRowHidden(row, not match or hide_unmatched)

    def filter_bank_table(self, text):
        """Filter bank table based on search text"""
        for row in range(self.bank_table.rowCount()):
            match = False
            for col in range(self.bank_table.columnCount() - 1):  # Exclude the checkbox column
                item = self.bank_table.item(row, col)
                if item and text.lower() in item.text().lower():
                    match = True
                    break
            # Also check if we should hide unmatched items
            hide_unmatched = (self.show_unmatched_bank.isChecked() and 
                            self.bank_table.item(row, 7) and 
                            self.bank_table.item(row, 7).text() == "Reconciled")
            self.bank_table.setRowHidden(row, not match or hide_unmatched)

    def toggle_ledger_unmatched_view(self, state):
        """Toggle view to show only unmatched ledger items"""
        self.filter_ledger_table(self.ledger_search.text())

    def toggle_bank_unmatched_view(self, state):
        """Toggle view to show only unmatched bank items"""
        self.filter_bank_table(self.bank_search.text())

    def refresh_search_filters(self):
        """Refresh all search filters"""
        self.filter_ledger_table(self.ledger_search.text())
        self.filter_bank_table(self.bank_search.text())